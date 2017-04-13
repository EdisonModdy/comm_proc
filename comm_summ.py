from openpyxl import load_workbook
import numpy as np
from openpyxl import Workbook
from pandas import DataFrame

sheetnum = 2
lengths = [946, 4155]
filename = 'comm_summ.xlsx'
wb = load_workbook(filename)

# find all the years and cases
years = set()
cases = set()

for i in range(sheetnum):
    sheetname = 'Batch' + str(i+1)
    length = lengths[i]
    ws = wb[sheetname]
    for row in range(3, length):
        case = ws['A'+str(row)].value
        cases.add(case)
        year = ws['C'+str(row)].value
        years.add(year)
cases = sorted(list(cases))
years = sorted(list(years))

summs = np.zeros((len(years), len(cases)))
summs = DataFrame(summs, index=years, columns=cases)
for i in range(sheetnum):
    sheetname = 'Batch' + str(i+1)
    length = lengths[i]
    ws = wb[sheetname]
    for row in range(3, length):
        case = ws['A'+str(row)].value
        year = ws['C'+str(row)].value
        summs[case][year] += 1.0

wbo = Workbook()
wso = wbo.active
filename = 'comm_nums.xlsx'
for i in range(len(cases)):
    wso[chr(ord('b')+i) + str(1)].value = cases[i]
for i in range(len(years)):
    wso['A'+str(2+i)].value = years[i]
for i in range(len(cases)):
    for j in range(len(years)):
        wso[chr(ord('b')+i) + str(2+j)].value = summs[cases[i]][years[j]]

wbo.save(filename)
